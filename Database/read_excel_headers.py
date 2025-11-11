# -*- coding: utf-8 -*-
import sys
import io
from openpyxl import load_workbook

# Set UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def read_excel_headers(file_path):
    """讀取Excel檔案的第一列（欄位名稱）"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        return headers
    except Exception as e:
        return [f"Error: {str(e)}"]

def write_to_file(output_file):
    """將Excel欄位輸出到文件"""
    files = {
        "19M_料號基本資料檔": r"C:\github\SBIR\Database\電笛_各M表\19M_料號基本資料檔(B010106)電笛.xlsx",
        "3M_單機資料檔": r"C:\github\SBIR\Database\電笛_各M表\3M_單機資料檔(3M)(B010103)電笛.xlsx",
        "20M_料號主要件號檔": r"C:\github\SBIR\Database\電笛_各M表\20M_料號主要件號檔(B010107)電笛.xlsx",
        "18M_單機零附件檔": r"C:\github\SBIR\Database\電笛_各M表\18M_單機零附件檔(B010105)電笛.xlsx",
        "16M_單機特性檔": r"C:\github\SBIR\Database\電笛_各M表\16M_單機特性檔(B010104)電笛.xlsx",
        "2M_單位構型檔": r"C:\github\SBIR\Database\電笛_各M表\2M_B010102單位構型檔_電笛.xlsx",
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        for name, filepath in files.items():
            f.write(f"\n{'='*60}\n")
            f.write(f"檔案: {name}\n")
            f.write(f"{'='*60}\n")
            headers = read_excel_headers(filepath)
            for idx, header in enumerate(headers, 1):
                f.write(f"{idx:2d}. {header}\n")

if __name__ == "__main__":
    output_file = r"C:\github\SBIR\Database\Excel欄位清單.txt"
    write_to_file(output_file)
    print(f"已將欄位清單寫入: {output_file}")
